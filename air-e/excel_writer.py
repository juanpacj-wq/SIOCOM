"""
Módulo para escribir los datos extraídos en un archivo Excel
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, Any, List
from datetime import datetime


class ExcelWriter:
    """Clase para escribir datos en archivos Excel con formato profesional"""
    
    def __init__(self, archivo_salida: str):
        """
        Inicializar el escritor de Excel
        
        Args:
            archivo_salida: Ruta del archivo Excel de salida
        """
        self.archivo_salida = archivo_salida
        self.workbook = Workbook()
        
        # Estilos predefinidos
        self.estilo_titulo = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        self.estilo_encabezado = Font(name='Arial', size=11, bold=True)
        self.estilo_normal = Font(name='Arial', size=10)
        
        self.relleno_titulo = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        self.relleno_encabezado = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        self.relleno_alternado = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        self.borde_fino = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.alineacion_centro = Alignment(horizontal='center', vertical='center')
        self.alineacion_izquierda = Alignment(horizontal='left', vertical='center')
        self.alineacion_derecha = Alignment(horizontal='right', vertical='center')
    
    def escribir_datos(self, datos: Dict[str, Any]):
        """
        Escribir todos los datos en diferentes hojas del Excel
        
        Args:
            datos: Diccionario con todos los datos extraídos
        """
        # Eliminar la hoja por defecto
        self.workbook.remove(self.workbook.active)
        
        # 1. Hoja de resumen general
        self._crear_hoja_resumen(datos)
        
        # 2. Hoja de información general
        if 'informacion_general' in datos and datos['informacion_general']:
            self._crear_hoja_informacion_general(datos['informacion_general'])
        
        # 3. Hoja de conceptos facturados
        if 'conceptos_facturados' in datos and datos['conceptos_facturados']:
            self._crear_hoja_conceptos(datos['conceptos_facturados'])
        
        # 4. Hoja de datos del cliente
        if 'datos_cliente' in datos and datos['datos_cliente']:
            self._crear_hoja_cliente(datos['datos_cliente'])
        
        # 5. Hoja de datos del operador
        if 'datos_operador' in datos and datos['datos_operador']:
            self._crear_hoja_operador(datos['datos_operador'])
        
        # 6. Hoja de periodo y fechas
        if 'periodo_facturacion' in datos and datos['periodo_facturacion']:
            self._crear_hoja_periodo(datos['periodo_facturacion'])
        
        # 7. Hoja de información de pago
        if 'informacion_pago' in datos and datos['informacion_pago']:
            self._crear_hoja_pago(datos['informacion_pago'])
        
        # 8. Hoja de todos los datos (tabla completa)
        self._crear_hoja_datos_completos(datos)
        
        # Guardar el archivo
        self.workbook.save(self.archivo_salida)
    
    def _crear_hoja_resumen(self, datos: Dict[str, Any]):
        """Crear hoja de resumen ejecutivo"""
        ws = self.workbook.create_sheet("Resumen")
        
        # Título
        ws.merge_cells('A1:D1')
        ws['A1'] = 'RESUMEN FACTURA AIR-E'
        ws['A1'].font = self.estilo_titulo
        ws['A1'].fill = self.relleno_titulo
        ws['A1'].alignment = self.alineacion_centro
        
        fila = 3
        
        # Información clave
        info_clave = []
        
        if 'informacion_general' in datos:
            info = datos['informacion_general']
            if 'nic' in info:
                info_clave.append(('NIC', info['nic']))
            if 'documento_equivalente' in info:
                info_clave.append(('Documento Equivalente', info['documento_equivalente']))
            if 'total_pagar' in info:
                info_clave.append(('Total a Pagar', f"${info['total_pagar']:,.0f}"))
        
        if 'datos_cliente' in datos:
            cliente = datos['datos_cliente']
            if 'nombre' in cliente:
                info_clave.append(('Cliente', cliente['nombre']))
            if 'nit' in cliente:
                info_clave.append(('NIT/CC', cliente['nit']))
        
        if 'periodo_facturacion' in datos:
            periodo = datos['periodo_facturacion']
            if 'fecha_vencimiento' in periodo:
                info_clave.append(('Fecha de Vencimiento', periodo['fecha_vencimiento']))
            if 'periodo_facturado' in periodo:
                info_clave.append(('Período Facturado', periodo['periodo_facturado']))
        
        # Escribir información clave
        for campo, valor in info_clave:
            ws[f'A{fila}'] = campo
            ws[f'A{fila}'].font = self.estilo_encabezado
            ws[f'B{fila}'] = valor
            ws[f'B{fila}'].font = self.estilo_normal
            
            # Aplicar bordes
            for col in ['A', 'B']:
                ws[f'{col}{fila}'].border = self.borde_fino
            
            fila += 1
        
        # Estadísticas
        fila += 2
        ws[f'A{fila}'] = 'ESTADÍSTICAS'
        ws[f'A{fila}'].font = self.estilo_encabezado
        ws[f'A{fila}'].fill = self.relleno_encabezado
        fila += 1
        
        # Contar elementos extraídos
        estadisticas = []
        for categoria, valores in datos.items():
            if valores:
                if isinstance(valores, dict):
                    estadisticas.append((categoria.replace('_', ' ').title(), f"{len(valores)} campos"))
                elif isinstance(valores, list):
                    estadisticas.append((categoria.replace('_', ' ').title(), f"{len(valores)} items"))
        
        for campo, valor in estadisticas:
            ws[f'A{fila}'] = campo
            ws[f'B{fila}'] = valor
            ws[f'A{fila}'].font = self.estilo_normal
            ws[f'B{fila}'].font = self.estilo_normal
            fila += 1
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
    
    def _crear_hoja_informacion_general(self, info: Dict[str, Any]):
        """Crear hoja con información general de la factura"""
        ws = self.workbook.create_sheet("Información General")
        
        # Título
        ws.merge_cells('A1:B1')
        ws['A1'] = 'INFORMACIÓN GENERAL DE LA FACTURA'
        ws['A1'].font = self.estilo_titulo
        ws['A1'].fill = self.relleno_titulo
        ws['A1'].alignment = self.alineacion_centro
        
        # Encabezados
        ws['A3'] = 'Campo'
        ws['B3'] = 'Valor'
        ws['A3'].font = self.estilo_encabezado
        ws['B3'].font = self.estilo_encabezado
        ws['A3'].fill = self.relleno_encabezado
        ws['B3'].fill = self.relleno_encabezado
        
        # Datos
        fila = 4
        for campo, valor in info.items():
            ws[f'A{fila}'] = campo.replace('_', ' ').title()
            ws[f'B{fila}'] = str(valor)
            
            # Formato para valores monetarios
            if 'total' in campo.lower() or 'valor' in campo.lower():
                if isinstance(valor, (int, float)):
                    ws[f'B{fila}'] = f"${valor:,.0f}"
            
            # Aplicar bordes
            ws[f'A{fila}'].border = self.borde_fino
            ws[f'B{fila}'].border = self.borde_fino
            
            # Color alternado
            if fila % 2 == 0:
                ws[f'A{fila}'].fill = self.relleno_alternado
                ws[f'B{fila}'].fill = self.relleno_alternado
            
            fila += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
    
    def _crear_hoja_conceptos(self, conceptos: List[Dict[str, Any]]):
        """Crear hoja con los conceptos facturados"""
        ws = self.workbook.create_sheet("Conceptos Facturados")
        
        # Título
        ws.merge_cells('A1:C1')
        ws['A1'] = 'CONCEPTOS FACTURADOS'
        ws['A1'].font = self.estilo_titulo
        ws['A1'].fill = self.relleno_titulo
        ws['A1'].alignment = self.alineacion_centro
        
        if conceptos:
            # Escribir encabezados
            fila_inicio = 3
            columnas = ['No.', 'Concepto', 'Valor']
            
            for col_idx, col_name in enumerate(columnas, 1):
                celda = ws.cell(row=fila_inicio, column=col_idx, value=col_name)
                celda.font = self.estilo_encabezado
                celda.fill = self.relleno_encabezado
                celda.border = self.borde_fino
                celda.alignment = self.alineacion_centro
            
            # Escribir datos
            for row_idx, concepto in enumerate(conceptos, start=fila_inicio + 1):
                # Número
                celda = ws.cell(row=row_idx, column=1, value=row_idx - fila_inicio)
                celda.border = self.borde_fino
                celda.alignment = self.alineacion_centro
                
                # Concepto
                celda = ws.cell(row=row_idx, column=2, value=concepto.get('concepto', ''))
                celda.border = self.borde_fino
                celda.alignment = self.alineacion_izquierda
                
                # Valor
                valor = concepto.get('valor', 0)
                celda = ws.cell(row=row_idx, column=3, value=valor)
                celda.border = self.borde_fino
                celda.number_format = '#,##0.00'
                celda.alignment = self.alineacion_derecha
                
                # Color alternado
                if row_idx % 2 == 0:
                    for col in range(1, 4):
                        ws.cell(row=row_idx, column=col).fill = self.relleno_alternado
            
            # Agregar fila de total
            fila_total = fila_inicio + len(conceptos) + 1
            ws[f'A{fila_total}'] = ''
            ws[f'B{fila_total}'] = 'TOTAL'
            ws[f'B{fila_total}'].font = self.estilo_encabezado
            ws[f'B{fila_total}'].alignment = self.alineacion_derecha
            
            # Calcular y mostrar total
            total = sum(c.get('valor', 0) for c in conceptos)
            celda_total = ws[f'C{fila_total}']
            celda_total.value = total
            celda_total.number_format = '#,##0.00'
            celda_total.font = self.estilo_encabezado
            celda_total.alignment = self.alineacion_derecha
            celda_total.border = self.borde_fino
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 20
    
    def _crear_hoja_cliente(self, datos: Dict[str, str]):
        """Crear hoja con datos del cliente"""
        ws = self.workbook.create_sheet("Datos Cliente")
        
        # Título
        ws.merge_cells('A1:B1')
        ws['A1'] = 'DATOS DEL CLIENTE'
        ws['A1'].font = self.estilo_titulo
        ws['A1'].fill = self.relleno_titulo
        ws['A1'].alignment = self.alineacion_centro
        
        # Datos
        fila = 3
        for campo, valor in datos.items():
            ws[f'A{fila}'] = campo.replace('_', ' ').title()
            ws[f'B{fila}'] = str(valor)
            ws[f'A{fila}'].font = self.estilo_encabezado
            ws[f'B{fila}'].font = self.estilo_normal
            ws[f'A{fila}'].border = self.borde_fino
            ws[f'B{fila}'].border = self.borde_fino
            
            # Color alternado
            if fila % 2 == 1:
                ws[f'A{fila}'].fill = self.relleno_alternado
                ws[f'B{fila}'].fill = self.relleno_alternado
            
            fila += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def _crear_hoja_operador(self, datos: Dict[str, str]):
        """Crear hoja con datos del operador de red"""
        ws = self.workbook.create_sheet("Operador de Red")
        
        # Título
        ws.merge_cells('A1:B1')
        ws['A1'] = 'INFORMACIÓN DEL OPERADOR DE RED'
        ws['A1'].font = self.estilo_titulo
        ws['A1'].fill = self.relleno_titulo
        ws['A1'].alignment = self.alineacion_centro
        
        # Datos
        fila = 3
        for campo, valor in datos.items():
            ws[f'A{fila}'] = campo.replace('_', ' ').title()
            ws[f'B{fila}'] = str(valor)
            ws[f'A{fila}'].font = self.estilo_encabezado
            ws[f'B{fila}'].font = self.estilo_normal
            ws[f'A{fila}'].border = self.borde_fino
            ws[f'B{fila}'].border = self.borde_fino
            fila += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def _crear_hoja_periodo(self, periodo: Dict[str, str]):
        """Crear hoja con información del periodo de facturación"""
        ws = self.workbook.create_sheet("Periodo Facturación")
        
        # Título
        ws.merge_cells('A1:B1')
        ws['A1'] = 'PERIODO DE FACTURACIÓN'
        ws['A1'].font = self.estilo_titulo
        ws['A1'].fill = self.relleno_titulo
        ws['A1'].alignment = self.alineacion_centro
        
        # Datos
        fila = 3
        for campo, valor in periodo.items():
            ws[f'A{fila}'] = campo.replace('_', ' ').title()
            ws[f'B{fila}'] = str(valor)
            ws[f'A{fila}'].font = self.estilo_encabezado
            ws[f'B{fila}'].font = self.estilo_normal
            ws[f'A{fila}'].border = self.borde_fino
            ws[f'B{fila}'].border = self.borde_fino
            
            # Resaltar fecha de vencimiento
            if 'vencimiento' in campo.lower():
                ws[f'B{fila}'].font = Font(name='Arial', size=10, bold=True, color='FF0000')
            
            fila += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
    
    def _crear_hoja_pago(self, pago: Dict[str, Any]):
        """Crear hoja con información de pago"""
        ws = self.workbook.create_sheet("Información de Pago")
        
        # Título
        ws.merge_cells('A1:B1')
        ws['A1'] = 'INFORMACIÓN DE PAGO'
        ws['A1'].font = self.estilo_titulo
        ws['A1'].fill = self.relleno_titulo
        ws['A1'].alignment = self.alineacion_centro
        
        # Datos
        fila = 3
        for campo, valor in pago.items():
            ws[f'A{fila}'] = campo.replace('_', ' ').title()
            
            # Formatear valores monetarios
            if isinstance(valor, (int, float)) and ('subtotal' in campo.lower() or 
                                                    'deuda' in campo.lower() or 
                                                    'saldo' in campo.lower() or
                                                    'facturacion' in campo.lower()):
                ws[f'B{fila}'] = f"${valor:,.2f}"
            else:
                ws[f'B{fila}'] = str(valor)
            
            ws[f'A{fila}'].font = self.estilo_encabezado
            ws[f'B{fila}'].font = self.estilo_normal
            ws[f'A{fila}'].border = self.borde_fino
            ws[f'B{fila}'].border = self.borde_fino
            
            # Resaltar saldo actual
            if 'saldo_actual' in campo.lower():
                ws[f'B{fila}'].font = Font(name='Arial', size=10, bold=True)
            
            fila += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
    
    def _crear_hoja_datos_completos(self, datos: Dict[str, Any]):
        """Crear hoja con todos los datos en formato tabla"""
        ws = self.workbook.create_sheet("Datos Completos")
        
        # Título
        ws.merge_cells('A1:C1')
        ws['A1'] = 'TODOS LOS DATOS EXTRAÍDOS'
        ws['A1'].font = self.estilo_titulo
        ws['A1'].fill = self.relleno_titulo
        ws['A1'].alignment = self.alineacion_centro
        
        # Encabezados
        ws['A3'] = 'Categoría'
        ws['B3'] = 'Campo'
        ws['C3'] = 'Valor'
        for col in ['A', 'B', 'C']:
            ws[f'{col}3'].font = self.estilo_encabezado
            ws[f'{col}3'].fill = self.relleno_encabezado
            ws[f'{col}3'].border = self.borde_fino
        
        # Datos
        fila = 4
        for categoria, valores in datos.items():
            if valores:
                if isinstance(valores, dict):
                    for campo, valor in valores.items():
                        ws[f'A{fila}'] = categoria.replace('_', ' ').title()
                        ws[f'B{fila}'] = campo.replace('_', ' ').title()
                        
                        # Manejar diferentes tipos de valores
                        if isinstance(valor, list):
                            ws[f'C{fila}'] = ', '.join(str(v) for v in valor)
                        elif isinstance(valor, (int, float)):
                            if 'total' in campo.lower() or 'valor' in campo.lower():
                                ws[f'C{fila}'] = f"${valor:,.2f}"
                            else:
                                ws[f'C{fila}'] = valor
                        else:
                            ws[f'C{fila}'] = str(valor)
                        
                        # Aplicar formato
                        for col in ['A', 'B', 'C']:
                            ws[f'{col}{fila}'].border = self.borde_fino
                            if fila % 2 == 0:
                                ws[f'{col}{fila}'].fill = self.relleno_alternado
                        
                        fila += 1
                
                elif isinstance(valores, list):
                    # Para listas (como conceptos_facturados)
                    for idx, item in enumerate(valores):
                        if isinstance(item, dict):
                            for campo, valor in item.items():
                                ws[f'A{fila}'] = f"{categoria.replace('_', ' ').title()} [{idx+1}]"
                                ws[f'B{fila}'] = campo.replace('_', ' ').title()
                                
                                if isinstance(valor, (int, float)) and 'valor' in campo.lower():
                                    ws[f'C{fila}'] = f"${valor:,.2f}"
                                else:
                                    ws[f'C{fila}'] = str(valor)
                                
                                for col in ['A', 'B', 'C']:
                                    ws[f'{col}{fila}'].border = self.borde_fino
                                    if fila % 2 == 0:
                                        ws[f'{col}{fila}'].fill = self.relleno_alternado
                                
                                fila += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 50
    
    def guardar_como_csv(self, datos: Dict[str, Any], archivo_csv: str = "factura_aire_datos.csv"):
        """
        Guardar los datos también en formato CSV
        
        Args:
            datos: Diccionario con los datos extraídos
            archivo_csv: Nombre del archivo CSV de salida
        """
        # Aplanar los datos para CSV
        filas = []
        
        for categoria, valores in datos.items():
            if valores:
                if isinstance(valores, dict):
                    for campo, valor in valores.items():
                        filas.append({
                            'Categoría': categoria,
                            'Campo': campo,
                            'Valor': valor
                        })
                elif isinstance(valores, list):
                    for idx, item in enumerate(valores):
                        if isinstance(item, dict):
                            for campo, valor in item.items():
                                filas.append({
                                    'Categoría': f"{categoria}_{idx+1}",
                                    'Campo': campo,
                                    'Valor': valor
                                })
        
        # Crear DataFrame y guardar
        if filas:
            df = pd.DataFrame(filas)
            df.to_csv(archivo_csv, index=False, encoding='utf-8-sig')
            print(f"  ✓ Datos también guardados en CSV: {archivo_csv}")